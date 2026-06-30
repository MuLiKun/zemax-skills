from __future__ import annotations

import argparse
import py_compile
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def _run(cmd: list[str], cwd: Path) -> None:
    print("$ " + " ".join(f'"{c}"' if " " in c else c for c in cmd))
    completed = subprocess.run(cmd, cwd=str(cwd))
    if completed.returncode != 0:
        raise SystemExit(completed.returncode)


def _run_expect_fail(cmd: list[str], cwd: Path) -> None:
    print("$ " + " ".join(f'"{c}"' if " " in c else c for c in cmd))
    completed = subprocess.run(cmd, cwd=str(cwd))
    if completed.returncode == 0:
        raise RuntimeError("负向检查失败：命令不应成功")


def _compile(paths: list[Path]) -> None:
    for path in paths:
        py_compile.compile(str(path), doraise=True)


def _copy_template(src: Path, dst: Path) -> Path:
    shutil.copy2(src, dst)
    return dst


def _set_run_param(path: Path, key: str, value) -> None:
    wb = load_workbook(path)
    ws = wb["输入_运行参数"]
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() == key:
            ws.cell(row=row, column=2, value=value)
            wb.save(path)
            return
    raise KeyError(f"未找到运行参数：{key}")


def _set_report_mf_line(path: Path, label: str, mf_line) -> None:
    wb = load_workbook(path)
    ws = wb["输入_REPORT"]
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value or "").strip() == label:
            ws.cell(row=row, column=3, value=mf_line)
            wb.save(path)
            return
    raise KeyError(f"未找到 REPORT：{label}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="一阶段基础检查：不连接 Zemax")
    parser.add_argument("--python", default=sys.executable,
                        help="用于运行 tol_run.py 的 Python 解释器，默认当前解释器")
    args = parser.parse_args(argv)

    script_dir = Path(__file__).resolve().parent
    py = Path(args.python).resolve()
    if not py.exists():
        raise FileNotFoundError(f"Python 解释器不存在：{py}")

    print("[1/10] Python 编译检查")
    py_files = [
        script_dir / "check_stage1.py",
        script_dir / "gui.py",
        script_dir / "tol_run.py",
        script_dir / "make_backup.py",
        *sorted((script_dir / "toltool").glob("*.py")),
    ]
    _compile(py_files)

    print("[2/10] 生成临时 Excel 模板")
    with tempfile.TemporaryDirectory(prefix="zemax_tol_stage1_") as tmp:
        tmp_dir = Path(tmp)
        template = tmp_dir / "tol_config_check.xlsx"
        dummy_zmx = tmp_dir / "dummy.zmx"
        dummy_zmx.write_text("SURF 0\nSURF 1\nSURF 2\nSURF 3\n", encoding="ascii")

        _run([str(py), "-u", "tol_run.py", "--init-template",
              "--config", str(template), "--overwrite"], cwd=script_dir)

        print("[3/10] validate-only 正向检查")
        _run([str(py), "-u", "tol_run.py", "--validate-only",
              "--zmx", str(dummy_zmx), "--config", str(template)], cwd=script_dir)

        print("[4/10] validate-only 负向检查：缺失 zmx")
        _run_expect_fail([str(py), "-u", "tol_run.py", "--validate-only",
                          "--zmx", str(tmp_dir / "missing.zmx"),
                          "--config", str(template)], cwd=script_dir)

        print("[5/10] validate-only 负向检查：缺失 Excel")
        _run_expect_fail([str(py), "-u", "tol_run.py", "--validate-only",
                          "--zmx", str(dummy_zmx),
                          "--config", str(tmp_dir / "missing.xlsx")], cwd=script_dir)

        print("[6/10] validate-only 负向检查：保存数量大于 MC 次数")
        bad_save_count = _copy_template(template, tmp_dir / "bad_save_count.xlsx")
        _set_run_param(bad_save_count, "蒙特卡洛次数", 5)
        _set_run_param(bad_save_count, "保存数量", 6)
        _run_expect_fail([str(py), "-u", "tol_run.py", "--validate-only",
                          "--zmx", str(dummy_zmx),
                          "--config", str(bad_save_count)], cwd=script_dir)

        print("[7/10] validate-only 负向检查：REPORT 指向不存在 MFE 行号")
        bad_report = _copy_template(template, tmp_dir / "bad_report.xlsx")
        _set_report_mf_line(bad_report, "SPOT_Hy0.0", 999)
        _run_expect_fail([str(py), "-u", "tol_run.py", "--validate-only",
                          "--zmx", str(dummy_zmx),
                          "--config", str(bad_report)], cwd=script_dir)

        print("[8/10] 标准模板模式 validate-only 正向检查")
        _run([str(py), "-u", "tol_run.py", "--validate-only", "--standard",
              "--zmx", str(dummy_zmx), "--outdir", str(tmp_dir),
              "--standard-template", "快速摸底", "--tolerance-level", "标准",
              "--num-runs", "5", "--num-to-save", "0"], cwd=script_dir)

        print("[9/10] 当前设置模式 validate-only 正向检查")
        _run([str(py), "-u", "tol_run.py", "--validate-only", "--current-settings",
              "--zmx", str(dummy_zmx), "--outdir", str(tmp_dir),
              "--num-runs", "5", "--num-to-save", "0"], cwd=script_dir)

        print("[10/10] 当前设置模式 validate-only 负向检查：缺失 zmx")
        _run_expect_fail([str(py), "-u", "tol_run.py", "--validate-only", "--current-settings",
                          "--zmx", str(tmp_dir / "missing_current.zmx"),
                          "--outdir", str(tmp_dir),
                          "--num-runs", "5", "--num-to-save", "0"], cwd=script_dir)

    print("基础检查通过。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
