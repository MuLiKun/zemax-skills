"""ztd_reader.py —— 读取公差分析 ZTD 文件并做分项统计（界面解耦）。

职责（需求文档 §7、skill §2e）：
- 打开 ToleranceDataViewer，加载 ZTD，读 MonteCarloData 矩阵。
- 矩阵列布局（probe 实测，05304_tol.ZTD 200x110）：
    col0      = 综合判据（评价函数总标量）
    col1..colN = N 个 REPORT 分项，顺序与 TSC 的 REPORT 完全一致
    其后各列 = 各单项公差灵敏度数据（本模块不解析）
- 对每个分项列统计：有效次数 N / 均值 / 标准差 / 最好 / 最差 / 2σ / 百分位。
- Cpk=1.33 规格限双边输出，并按 REPORT 方向标记用户更关注的一侧。
- 标签优先用调用方传入的 REPORT 标签；否则从 Summary「相对评估脚本」段解析。

读取陷阱（skill §2e，务必遵守）：
- 读结果前工具一次只能开一个；本模块自行 OpenToleranceDataViewer。
- 矩阵 Rows 对大矩阵可能惰性返回 0，不依赖它做边界；硬编码 nrow=num_runs。
- 不在主循环前对 Values 做任何预读（GetValueAt warmup / Data / TotalLength 轮询）。
- Values 后直接进 GetValueAt(i,j) 双重循环、行优先连续读满全部列、循环内禁 I/O。
"""

from __future__ import annotations

import math
import os
import re
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class ItemStat:
    label: str
    col: int
    n: int
    mean: float
    std: float
    best: float
    worst: float
    two_sigma: float
    nominal: float = float("nan")
    median: float = float("nan")
    p1: float = float("nan")
    p5: float = float("nan")
    p95: float = float("nan")
    p99: float = float("nan")
    direction: str = ""
    unit: str = ""
    lsl_cpk133: float = float("nan")
    usl_cpk133: float = float("nan")
    samples: list = field(default_factory=list)


@dataclass
class ZtdResult:
    succeeded: bool
    ztd_path: str
    num_runs: int
    num_cols: int
    items: list = field(default_factory=list)
    summary: str = ""
    message: str = ""


_SCRIPT_LINE = re.compile(r"^\s*(.+?)\s*=\s*\t?\s*([-+0-9.eE]+)\s*$")
_REPORT_LINE = re.compile(r'^\s*REPORT\s+"([^"]+)"\s+([0-9]+)\s*$', re.IGNORECASE)
_CPK_TARGET = 1.33


def _clean_label(value) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in ("none", "nan"):
        return ""
    return re.sub(r"\s+", " ", text)


def _try_call(obj, name: str, *args):
    try:
        member = getattr(obj, name)
    except Exception:
        return None
    try:
        if callable(member):
            return member(*args)
        if args:
            return None
        return member
    except Exception:
        return None


def _item_at(seq, index: int):
    for idx in (index, index + 1):
        for name in ("GetValueAt", "GetValue", "GetKey", "GetItem", "Item", "get_Item"):
            value = _try_call(seq, name, idx)
            if value is not None:
                return value
        try:
            return seq[idx]
        except Exception:
            pass
    return None


def _label_from_container(container, index: int) -> str:
    if container is None:
        return ""
    for idx in (index, index + 1):
        for name in (
                "GetColumnHeader", "GetColumnHeaderAt",
                "GetColumnLabel", "GetColumnLabelAt",
                "GetColumnName", "GetColumnNameAt",
                "GetColumnTitle", "GetColumnTitleAt",
                "GetColumnDescription", "GetHeaderAt", "GetLabelAt",
                "GetNameAt", "GetTitleAt", "GetKeyAt"):
            label = _clean_label(_try_call(container, name, idx))
            if label:
                return label
    for name in (
            "ColumnHeaders", "ColumnLabels", "ColumnNames", "ColumnTitles",
            "ColumnDescriptions", "Headers", "Labels", "Names", "Titles", "Keys"):
        labels = _try_call(container, name)
        label = _clean_label(_item_at(labels, index)) if labels is not None else ""
        if label:
            return label
    return ""


def _ztd_column_labels(mc, ncol: int):
    labels = []
    containers = [mc]
    for i in range(ncol):
        label = ""
        for container in containers:
            label = _label_from_container(container, i)
            if label:
                break
        labels.append(label)
    return labels


def _candidate_tsc_dirs(ztd_path: str, zos_system=None) -> list[str]:
    ztd_dir = os.path.dirname(os.path.abspath(ztd_path))
    dirs = [ztd_dir]
    try:
        mf_dir = str(zos_system.MFE.MeritFunctionDirectory)
        data_dir = os.path.dirname(mf_dir.rstrip("\\/"))
        dirs.append(os.path.join(data_dir, "Tolerance"))
    except Exception:
        pass
    for env_name in ("USERPROFILE", "HOME"):
        home = os.environ.get(env_name)
        if home:
            dirs.append(os.path.join(home, "Documents", "Zemax", "Tolerance"))
            dirs.append(os.path.join(home, "文档", "Zemax", "Tolerance"))
    for env_name in ("OneDrive", "OneDriveCommercial", "OneDriveConsumer"):
        root = os.environ.get(env_name)
        if root:
            dirs.append(os.path.join(root, "Documents", "Zemax", "Tolerance"))
            dirs.append(os.path.join(root, "文档", "Zemax", "Tolerance"))
    out = []
    seen = set()
    for d in dirs:
        if not d:
            continue
        path = os.path.abspath(d)
        key = os.path.normcase(path)
        if key not in seen:
            seen.add(key)
            out.append(path)
    return out


def _candidate_tsc_paths(ztd_path: str, zos_system=None) -> list[str]:
    ztd_abs = os.path.abspath(ztd_path)
    base = os.path.splitext(os.path.basename(ztd_abs))[0]
    names = [base + ".TSC", base + ".tsc", base + "_api.TSC", base + "_api.tsc"]
    out = []
    seen = set()
    for d in _candidate_tsc_dirs(ztd_path, zos_system):
        for name in names:
            path = os.path.abspath(os.path.join(d, name))
            key = os.path.normcase(path)
            if key not in seen:
                seen.add(key)
                out.append(path)
    return out


def _read_text_with_fallback(path: str) -> str:
    encodings = ["utf-8-sig", "utf-8", "mbcs", "latin-1"]
    last_err = None
    for encoding in encodings:
        try:
            with open(path, "r", encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError as e:
            last_err = e
        except LookupError:
            pass
    if last_err:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    with open(path, "r") as f:
        return f.read()


def _parse_tsc_report_labels(tsc_path: str) -> list[str]:
    labels = []
    text = _read_text_with_fallback(tsc_path)
    for raw in text.splitlines():
        m = _REPORT_LINE.match(raw)
        if not m:
            continue
        label = _clean_label(m.group(1))
        mf_line = _clean_label(m.group(2))
        if label and mf_line:
            labels.append(f"{label} ({mf_line})")
    return labels


def _tsc_labels_for_ztd(ztd_path: str, zos_system=None) -> tuple[list[str], str, list[str]]:
    candidates = _candidate_tsc_paths(ztd_path, zos_system)
    for path in candidates:
        if not os.path.isfile(path):
            continue
        labels = _parse_tsc_report_labels(path)
        if labels:
            return labels, path, candidates
    return [], "", candidates


def parse_summary_labels(summary: str):
    """从 Summary 的「相对评估脚本」段解析 (标签, 标称值) 列表，按出现顺序。"""
    out = []
    in_block = False
    for raw in str(summary).splitlines():
        line = raw.rstrip()
        if "相对评估脚本" in line or "Relative" in line:
            in_block = True
            continue
        if in_block:
            m = _SCRIPT_LINE.match(line)
            if m:
                label = m.group(1).strip()
                try:
                    nominal = float(m.group(2))
                except ValueError:
                    nominal = float("nan")
                out.append((label, nominal))
            elif line.strip() == "" and out:
                break
    return out


def _percentile(sorted_vals, pct: float) -> float:
    if not sorted_vals:
        return float("nan")
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    pos = (len(sorted_vals) - 1) * pct / 100.0
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return sorted_vals[lo]
    frac = pos - lo
    return sorted_vals[lo] * (1.0 - frac) + sorted_vals[hi] * frac


def _cpk_limits(mean: float, std: float, direction: str, cpk: float = _CPK_TARGET):
    if math.isnan(mean) or math.isnan(std):
        return float("nan"), float("nan")
    span = 3.0 * cpk * std
    return mean - span, mean + span


def _direction_focus(direction: str) -> str:
    d = str(direction or "").strip().lower()
    if "小" in d or d in ("min", "smaller", "lower", "less", "小于"):
        return "upper"
    if "大" in d or d in ("max", "larger", "higher", "greater", "大于"):
        return "lower"
    return ""


def _stats(values, direction: str = ""):
    vals = [v for v in values if v is not None and not math.isnan(v)]
    n = len(vals)
    if n == 0:
        nan = float("nan")
        return 0, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan, nan
    mean = sum(vals) / n
    if n > 1:
        var = sum((v - mean) ** 2 for v in vals) / (n - 1)
        std = math.sqrt(var)
    else:
        std = 0.0
    ordered = sorted(vals)
    lsl, usl = _cpk_limits(mean, std, direction)
    return (
        n,
        mean,
        std,
        min(vals),
        max(vals),
        2.0 * std,
        _percentile(ordered, 50),
        _percentile(ordered, 1),
        _percentile(ordered, 5),
        _percentile(ordered, 95),
        _percentile(ordered, 99),
        lsl,
        usl,
    )


def read_ztd(zos_system, ztd_path: str, num_runs: int,
             report_labels=None, num_items=None, report_meta=None) -> ZtdResult:
    """打开 ZTD，统计 col1..colN 各分项。"""
    dv = zos_system.Tools.OpenToleranceDataViewer()
    try:
        dv.FileName = ztd_path
        ok = bool(dv.RunAndWaitForCompletion())
        if not ok or not dv.Succeeded:
            return ZtdResult(False, ztd_path, num_runs, 0,
                             message=str(dv.ErrorMessage) or "加载 ZTD 失败")

        summary = ""
        try:
            summary = str(dv.Summary)
        except Exception:
            pass

        parsed = parse_summary_labels(summary)
        meta = list(report_meta or [])
        config_labels = []
        if report_labels:
            config_labels = [_clean_label(x) for x in report_labels]
        elif meta:
            config_labels = [_clean_label(m.get("标签")) for m in meta]
        parsed_labels = [p[0] for p in parsed]
        tsc_labels, tsc_path, tsc_candidates = _tsc_labels_for_ztd(ztd_path, zos_system)

        nominals = {p[0]: p[1] for p in parsed}

        mc = dv.MonteCarloData
        v = mc.Values
        ncol = int(v.Cols)
        ztd_labels = _ztd_column_labels(mc, ncol)

        nrow = int(num_runs)
        if num_items is not None:
            data_cols = list(range(0, min(int(num_items), ncol)))
        else:
            data_cols = list(range(0, ncol))

        buckets = {c: [] for c in data_cols}
        actual_rows = 0
        row_limited = False
        for i in range(nrow):
            row_values = []
            try:
                for j in data_cols:
                    row_values.append(v.GetValueAt(i, j))
            except Exception as e:
                if i > 0 and "IndexOutOfRange" in type(e).__name__:
                    row_limited = True
                    break
                if i > 0 and "索引超出了数组界限" in str(e):
                    row_limited = True
                    break
                raise
            for j, value in zip(data_cols, row_values):
                buckets[j].append(value)
            actual_rows += 1

        if data_cols and actual_rows <= 0:
            return ZtdResult(False, ztd_path, num_runs, ncol,
                             message="ZTD 中未能读取到 Monte Carlo 数据行。")

        items = []
        tsc_used = 0
        for idx, c in enumerate(data_cols):
            label = ""
            if c < len(ztd_labels):
                label = _clean_label(ztd_labels[c])
            fallback_idx = c - 1 if c > 0 else -1
            if not label and c == 0 and tsc_labels:
                label = "自定义脚本"
            if not label and 0 <= fallback_idx < len(tsc_labels):
                label = _clean_label(tsc_labels[fallback_idx])
                if label:
                    tsc_used += 1
            if not label and 0 <= fallback_idx < len(parsed_labels):
                label = _clean_label(parsed_labels[fallback_idx])
            if not label and 0 <= fallback_idx < len(config_labels):
                label = _clean_label(config_labels[fallback_idx])
            if not label:
                label = f"col{c}"
            direction = ""
            unit = ""
            meta_idx = c - 1 if c > 0 else -1
            if 0 <= meta_idx < len(meta):
                direction = str(meta[meta_idx].get("方向") or "").strip()
                unit = str(meta[meta_idx].get("单位") or "").strip()
            stat = _stats(buckets[c], direction=direction)
            (n, mean, std, vmin, vmax, two_s, median, p1, p5, p95, p99,
             lsl, usl) = stat
            items.append(ItemStat(
                label=label, col=c, n=n, mean=mean, std=std,
                best=vmin, worst=vmax, two_sigma=two_s,
                nominal=nominals.get(label, float("nan")),
                median=median, p1=p1, p5=p5, p95=p95, p99=p99,
                direction=direction, unit=unit,
                lsl_cpk133=lsl, usl_cpk133=usl,
                samples=list(buckets[c]),
            ))

        messages = []
        if row_limited:
            messages.append(
                f"配置蒙特卡洛次数为 {num_runs}，ZTD 实际可读取 {actual_rows} 行。")
        if tsc_path:
            messages.append(
                f"已从 TSC 反推 {tsc_used}/{len(tsc_labels)} 个 REPORT 表头：{tsc_path}")
        else:
            searched = "；".join(tsc_candidates)
            messages.append(
                f"未搜索到同名 TSC，无法反推 REPORT 表头，Excel 将按 colN 输出。已搜索：{searched}")
        unresolved = sum(1 for it in items if str(it.label).startswith("col"))
        if unresolved:
            messages.append(
                f"仍有 {unresolved} 个 ZTD 矩阵列未读取到列标题，已使用 colN 兜底。")

        return ZtdResult(
            succeeded=True, ztd_path=ztd_path, num_runs=actual_rows,
            num_cols=ncol, items=items, summary=summary,
            message="；".join(messages),
        )
    finally:
        try:
            dv.Close()
        except Exception:
            pass


def format_table(result: ZtdResult) -> str:
    """把 ZtdResult 渲染成对齐文本表，便于命令行汇报。"""
    if not result.succeeded:
        return f"读取失败: {result.message}"
    lines = [
        f"ZTD: {result.ztd_path}",
        f"蒙特卡洛次数: {result.num_runs}  矩阵列数: {result.num_cols}  "
        f"分项数: {len(result.items)}",
        "",
        f"{'分项':<14}{'N':>5}{'方向':>6}{'标称':>14}{'均值':>14}"
        f"{'标准差':>14}{'P50':>14}{'P5':>14}{'P95':>14}"
        f"{'Cpk1.33下限':>16}{'Cpk1.33上限':>16}",
    ]
    for it in result.items:
        lines.append(
            f"{it.label:<14}{it.n:>5}{it.direction:>6}{it.nominal:>14.6g}"
            f"{it.mean:>14.6g}{it.std:>14.6g}{it.median:>14.6g}"
            f"{it.p5:>14.6g}{it.p95:>14.6g}"
            f"{it.lsl_cpk133:>16.6g}{it.usl_cpk133:>16.6g}"
        )
    return "\n".join(lines)


def export_excel(result: ZtdResult, path: str) -> str:
    if not result.succeeded:
        raise RuntimeError(result.message or "ZTD 读取失败，无法导出统计 Excel")
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "ZTD统计"
    max_samples = max([len(it.samples) for it in result.items] or [result.num_runs])
    headers = ["项目"] + [it.label for it in result.items]
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    stat_fill = PatternFill("solid", fgColor="FFF2CC")
    focus_fill = PatternFill("solid", fgColor="FFFF00")
    data_fill = PatternFill("solid", fgColor="E2F0D9")
    font = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Cpk1.33下限", [it.lsl_cpk133 for it in result.items]),
        ("Cpk1.33上限", [it.usl_cpk133 for it in result.items]),
        ("最大值", [it.worst for it in result.items]),
        ("最小值", [it.best for it in result.items]),
        ("平均值", [it.mean for it in result.items]),
        ("标准差", [it.std for it in result.items]),
        ("中位数", [it.median for it in result.items]),
    ]
    for i in range(max_samples):
        values = [it.samples[i] if i < len(it.samples) else None
                  for it in result.items]
        rows.append((f"MC{i + 1:03d}", values))

    for r, (name, values) in enumerate(rows, start=2):
        name_cell = ws.cell(row=r, column=1, value=name)
        if r <= 8:
            name_cell.fill = stat_fill
            name_cell.font = font
        else:
            name_cell.fill = data_fill
        for c, value in enumerate(values, start=2):
            cell = ws.cell(row=r, column=c, value=value)
            item = result.items[c - 2] if c - 2 < len(result.items) else None
            focus = _direction_focus(item.direction) if item else ""
            if (name == "Cpk1.33下限" and focus == "lower") or (
                    name == "Cpk1.33上限" and focus == "upper"):
                cell.fill = focus_fill
                cell.font = font

    ws.freeze_panes = "B9"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 18
    for c in range(2, len(headers) + 1):
        letter = ws.cell(row=1, column=c).column_letter
        header = ws.cell(row=1, column=c).value
        ws.column_dimensions[letter].width = max(12, min(28, len(str(header)) + 2))

    ws2 = wb.create_sheet("说明")
    lines = [
        ["ZTD文件", result.ztd_path],
        ["统计结构", "列为 REPORT 指标；前 7 行为统计项；后续 MC001... 为每轮 Monte Carlo 原始数据。"],
        ["统计项", "Cpk1.33下限、Cpk1.33上限、最大值、最小值、平均值、标准差、中位数"],
        ["蒙特卡洛次数", result.num_runs],
        ["矩阵列数", result.num_cols],
        ["Cpk目标", _CPK_TARGET],
        ["Cpk1.33下限", "LSL = 均值 - 3*Cpk*标准差"],
        ["Cpk1.33上限", "USL = 均值 + 3*Cpk*标准差"],
        ["方向", "方向保留用于标识越小越好/越大越好；统计 Excel 始终输出双边上下限。"],
        ["黄色标记", "方向为越小越好时标黄 Cpk1.33上限；方向为越大越好时标黄 Cpk1.33下限。"],
    ]
    if result.message:
        lines.append(["提示", result.message])
    for r, row in enumerate(lines, start=1):
        for c, value in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=value)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 96
    wb.save(path)
    return path
