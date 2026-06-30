"""current_settings.py —— 复用 Zemax 当前 TDE/MFE 的配置生成。

当前设置模式不从 Excel 或标准模板重建 TDE/MFE，而是在工作副本打开后：
- 保留镜头文件中已有 TDE；
- 保存当前 MFE 为 .MF；
- 扫描当前 MFE 有效行，自动生成 REPORT/TSC；
- 运行参数仍写入 Config，便于后续统计与 used_excel.xlsx 复核。
"""

from __future__ import annotations

import os
import re

from openpyxl import load_workbook

from . import excel_io
from .field_mapping import build_field_items, report_label as _field_report_label

_FIELD_SMALL_OPS = {"RSCE", "RWCE", "GENC"}
_FIELD_LARGE_OPS = {"GMTT", "GMTS", "GMTA", "MTFT", "MTFS", "MTFA"}
_PARAM_COLS = [f"Param{i}" for i in range(1, 9)]
_OP_RE = re.compile(r"[A-Z]{3,4}")


def _enum_name(value) -> str:
    text = str(value or "").strip()
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    text = text.upper()
    match = _OP_RE.search(text)
    return match.group(0) if match else text


def _cell_value(row, merit_column, name: str):
    try:
        col = getattr(merit_column, name)
    except AttributeError:
        return ""
    try:
        return row.GetOperandCell(col).Value
    except Exception:
        return ""


def _row_op(row) -> str:
    for attr in ("Type", "OperandType"):
        try:
            op = _enum_name(getattr(row, attr))
            if op:
                return op
        except Exception:
            pass
    return ""


def _row_comment(row) -> str:
    for attr in ("Comment", "CommentText"):
        try:
            text = str(getattr(row, attr) or "").strip()
            if text:
                return text
        except Exception:
            pass
    return ""


def _direction(op: str) -> str:
    if op in _FIELD_LARGE_OPS:
        return "大"
    if op in _FIELD_SMALL_OPS:
        return "小"
    return ""


def _unit(op: str) -> str:
    if op in {"RSCE", "RWCE"}:
        return "mm"
    if op == "GENC":
        return "um"
    if op in _FIELD_LARGE_OPS:
        return "-"
    return ""


def _field_normalized_by_no(fields_by_no: dict[int, float], value) -> float | None:
    try:
        field_no = int(float(value))
    except (TypeError, ValueError):
        return None
    return fields_by_no.get(field_no)


def _report_label(op: str, line_no: int, item: dict,
                  fields_by_no: dict[int, float]) -> str:
    target = None
    if op in _FIELD_LARGE_OPS or op == "GENC":
        target = _field_normalized_by_no(fields_by_no, item.get("Param3"))
    elif op in {"RSCE", "RWCE"}:
        try:
            target = float(item.get("Param4"))
        except (TypeError, ValueError):
            target = None
    if target is not None:
        return f"{op}_{_field_report_label(target)}"
    return f"{op}_{line_no}"


def _current_field_normalized_map(zos_system) -> dict[int, float]:
    fields = zos_system.SystemData.Fields
    items = build_field_items([
        (float(fields.GetField(i).X), float(fields.GetField(i).Y))
        for i in range(1, fields.NumberOfFields + 1)
    ])
    return {item.field_no: item.normalized for item in items}


def read_current_mfe(zos_system) -> tuple[list[dict], list[dict]]:
    """读取当前 MFE 有效行，返回 (mfe_rows, report_rows)。"""
    import ZOSAPI

    mfe = zos_system.MFE
    merit_column = ZOSAPI.Editors.MFE.MeritColumn
    fields_by_no = _current_field_normalized_map(zos_system)
    rows: list[dict] = []
    report: list[dict] = []

    for line_no in range(1, int(mfe.NumberOfOperands) + 1):
        row = mfe.GetOperandAt(line_no)
        op = _row_op(row)
        if not op or op == "BLNK":
            continue
        comment = _row_comment(row)
        item = {
            "行号": line_no,
            "操作数": op,
            "目标": getattr(row, "Target", ""),
            "权重": getattr(row, "Weight", ""),
            "注释": comment or f"当前MFE第{line_no}行",
            "目标归一化视场": "",
            "视场映射说明": "当前设置模式读取",
            "归一化视场": "",
        }
        for name in _PARAM_COLS:
            item[name] = _cell_value(row, merit_column, name)
        if op in _FIELD_LARGE_OPS or op == "GENC":
            target = _field_normalized_by_no(fields_by_no, item.get("Param3"))
        elif op in {"RSCE", "RWCE"}:
            try:
                target = float(item.get("Param4"))
            except (TypeError, ValueError):
                target = None
        else:
            target = None
        if target is not None:
            item["目标归一化视场"] = target
            item["归一化视场"] = target
            item["视场映射说明"] = "当前设置模式按视场号反推"
        rows.append(item)
        report.append({
            "启用": "Y",
            "标签": _report_label(op, line_no, item, fields_by_no),
            "MF行号": line_no,
            "方向": _direction(op),
            "单位": _unit(op),
        })

    if not rows:
        raise ValueError("当前 MFE 中没有可用于 REPORT 的有效操作数。")
    return rows, report


def tde_has_comp(zos_system) -> bool:
    """检测当前 TDE 是否含有 COMP 补偿器操作数。"""
    try:
        tde = zos_system.TDE
        n = int(tde.NumberOfOperands)
        for i in range(1, n + 1):
            r = tde.GetOperandAt(i)
            try:
                type_name = _enum_name(r.Type)
            except Exception:
                type_name = ""
            if type_name == "COMP":
                return True
        return False
    except Exception:
        return False


def detect_comp_freq_from_mfe(cfg: excel_io.Config) -> float | None:
    """从当前 MFE 配置中提取 MTF 类操作数的空间频率（Param4）。"""
    freqs: list[float] = []
    for row in cfg.mfe:
        op = str(row.get("操作数") or "").strip().upper()
        if op in {"GMTT", "GMTS", "GMTA", "MTFT", "MTFS", "MTFA"}:
            try:
                freqs.append(float(row.get("Param4") or 0))
            except (TypeError, ValueError):
                pass
    return freqs[0] if freqs else None


def summarize_config(cfg: excel_io.Config, limit: int = 8) -> list[str]:
    lines = [
        f"当前设置模式：MFE 有效行 {len(cfg.mfe)}，自动 REPORT {len(cfg.report)}。"
    ]
    for row, rep in zip(cfg.mfe[:limit], cfg.report[:limit]):
        line = row.get("行号")
        op = row.get("操作数")
        target = row.get("目标")
        weight = row.get("权重")
        label = rep.get("标签")
        lines.append(
            f"  MFE[{line}] {op} 目标={target} 权重={weight} → REPORT {label}")
    remain = len(cfg.mfe) - limit
    if remain > 0:
        lines.append(f"  ... 其余 {remain} 行已写入 used_excel.xlsx")
    return lines


def build_config_from_current_mfe(zos_system, num_runs: int = 20,
                                  num_to_save: int = 0,
                                  comp_mode: str = "无",
                                  save_worst_best: bool = False) -> excel_io.Config:
    mfe, report = read_current_mfe(zos_system)
    return excel_io.Config(
        tol_wizard=[],
        tol_detail=[],
        mfe=mfe,
        report=report,
        run_params={
            "分析模式": "当前设置",
            "蒙特卡洛次数": int(num_runs),
            "保存数量": int(num_to_save),
            "统计分布": "正态",
            "补偿器模式": comp_mode,
            "TSC优化周期": 4,
            "中心波长号": "",
            "后焦补偿面": "",
            "补偿Min": "",
            "补偿Max": "",
            "补偿线对": "",
            "保存TSC": "Y",
            "保存WorstCase": "Y" if save_worst_best else "N",
            "保存BestCase": "Y" if save_worst_best else "N",
            "输出统计Excel": "Y",
            "输出直方图": "N",
            "启用视场映射": "N",
            "视场插入策略": "禁用",
            "视场匹配阈值": 0.05,
            "目标归一化视场": "",
            "目标视场来源策略": "自动推断",
        },
    )


def _rewrite_sheet(ws, header: list[str], rows: list[dict]) -> None:
    ws.delete_rows(2, ws.max_row)
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(header, start=1):
            ws.cell(row=r, column=c, value=row.get(key))


def write_config_excel(path: str, cfg: excel_io.Config, overwrite: bool = True) -> str:
    if os.path.exists(path) and not overwrite:
        raise FileExistsError(path)
    excel_io.generate_template(path, overwrite=True)
    wb = load_workbook(path)
    _rewrite_sheet(wb["输入_公差向导"], excel_io._TOL_WIZARD_HDR, cfg.tol_wizard)
    _rewrite_sheet(wb["输入_公差明细"], excel_io._TOL_DETAIL_HDR, cfg.tol_detail)
    _rewrite_sheet(wb["输入_评价函数"], excel_io._MFE_HDR, cfg.mfe)
    _rewrite_sheet(wb["输入_REPORT"], excel_io._REPORT_HDR, cfg.report)
    run_rows = [{"参数键": k, "值": v, "备注": "当前设置模式生成"} for k, v in cfg.run_params.items()]
    _rewrite_sheet(wb["输入_运行参数"], excel_io._RUN_HDR, run_rows)
    wb.save(path)
    return path
