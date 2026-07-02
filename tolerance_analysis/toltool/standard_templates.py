"""standard_templates.py —— 普通标准模板模式的内置配置生成。

模板内容集中在本文件顶部的数据区，后续正式标准确定后优先改
_LEVEL_VALUES / _TEMPLATES，不需要改 Excel 写出与主流程接入逻辑。
"""

from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass

from openpyxl import load_workbook

from . import excel_io


@dataclass(frozen=True)
class OperandSpec:
    label: str
    op: str
    params: dict
    direction: str
    unit: str
    target: float = 0
    weight: float = 1


@dataclass(frozen=True)
class TemplateSpec:
    name: str
    description: str
    target_fields: tuple[float, ...]
    operands: tuple[OperandSpec, ...]


_LEVEL_VALUES = {
    "宽松": {
        "半径": 3,
        "厚度": 0.05,
        "面倾斜": 0.05,
        "元件偏心": 0.05,
        "元件倾斜": 0.3,
        "面不规则": 1,
        "折射率": 0.0005,
        "阿贝%": 1,
    },
    "标准": {
        "半径": 3,
        "厚度": 0.03,
        "面倾斜": 0.05,
        "元件偏心": 0.03,
        "元件倾斜": 0.2,
        "面不规则": 1,
        "折射率": 0.0005,
        "阿贝%": 1,
    },
    "严格": {
        "半径": 3,
        "厚度": 0.02,
        "面倾斜": 0.025,
        "元件偏心": 0.02,
        "元件倾斜": 0.1,
        "面不规则": 1,
        "折射率": 0.0005,
        "阿贝%": 1,
    },
}


def _spot(label: str, field: float) -> OperandSpec:
    return OperandSpec(
        label=label,
        op="RSCE",
        params={
            "Param1": 3,
            "Param2": "{center_wave}",
            "Param3": 0,
            "Param4": field,
            "目标归一化视场": field,
        },
        direction="小",
        unit="mm",
    )


def _genc(label: str, field_no: int, target_field: float) -> OperandSpec:
    return OperandSpec(
        label=label,
        op="GENC",
        params={
            "Param1": 3,
            "Param2": "{center_wave}",
            "Param3": field_no,
            "Param4": 0.95,
            "Param5": 1,
            "Param6": 0,
            "Param7": 0,
            "目标归一化视场": target_field,
        },
        direction="小",
        unit="um",
    )


def _mtf(label: str, op: str, field_no: int, freq: float,
         target_field: float) -> OperandSpec:
    return OperandSpec(
        label=label,
        op=op,
        params={
            "Param1": 3,
            "Param2": "{center_wave}",
            "Param3": field_no,
            "Param4": freq,
            "Param5": 0,
            "Param6": 0,
            "目标归一化视场": target_field,
        },
        direction="大",
        unit="-",
    )


STANDARD_TARGET_FIELDS = (0, 0.5, 0.9, -0.9)
FULL_TARGET_FIELDS = (0, -0.25, 0.25, -0.5, 0.5, -0.7, 0.7, -0.9, 0.9, -1, 1)
PRODUCT_TYPES = ("RX", "TX")
PRODUCT_DESCRIPTIONS = {
    "RX": "RX：接收端标准模板。",
    "TX": "TX：发射端标准模板，当前暂复用 RX，后续可单独维护。",
}
DEFAULT_PRODUCT_TYPE = "RX"
DEFAULT_TEMPLATE_NAME = "标准分析"


def _field_label(value: float) -> str:
    if abs(float(value)) < 1e-12:
        return "F0"
    return f"F{float(value):g}"


def _spot_operands(fields: tuple[float, ...]) -> tuple[OperandSpec, ...]:
    return tuple(_spot(f"SPOT_{_field_label(field)}", field) for field in fields)


def _genc_operands(fields: tuple[float, ...]) -> tuple[OperandSpec, ...]:
    return tuple(
        _genc(f"GENC95_{_field_label(field)}", 1, field)
        for field in fields
    )


def _mtf_operands(fields: tuple[float, ...]) -> tuple[OperandSpec, ...]:
    rows: list[OperandSpec] = []
    for field in fields:
        label = _field_label(field)
        rows.append(_mtf(f"GMTFT_{label}", "GMTT", 1, 34, field))
        rows.append(_mtf(f"GMTFS_{label}", "GMTS", 1, 34, field))
    return tuple(rows)


_RX_TEMPLATES = {
    "标准分析": TemplateSpec(
        name="标准分析",
        description="标准分析：使用 0、0.5、0.9、-0.9 目标视场；包含点列评价，并对 ±0.9 边缘视场增加 GENC/MTF 评价。",
        target_fields=STANDARD_TARGET_FIELDS,
        operands=(
            *_spot_operands(STANDARD_TARGET_FIELDS),
            *_genc_operands((0.9, -0.9)),
            *_mtf_operands((0.9, -0.9)),
        ),
    ),
    "完整视场分析": TemplateSpec(
        name="完整视场分析",
        description="完整视场分析：使用 0、±0.25、±0.5、±0.7、±0.9、±1 全视场序列；包含全视场 SPOT、GENC、GMTT、GMTS 评价。",
        target_fields=FULL_TARGET_FIELDS,
        operands=(
            *_spot_operands(FULL_TARGET_FIELDS),
            *_genc_operands(FULL_TARGET_FIELDS),
            *_mtf_operands(FULL_TARGET_FIELDS),
        ),
    ),
}
_PRODUCT_TEMPLATES = {
    "RX": _RX_TEMPLATES,
    "TX": _RX_TEMPLATES,
}


TEMPLATE_NAMES = tuple(_RX_TEMPLATES.keys())
LEVEL_NAMES = tuple(_LEVEL_VALUES.keys())


def product_types() -> tuple[str, ...]:
    return PRODUCT_TYPES


def product_description(product_type: str = DEFAULT_PRODUCT_TYPE) -> str:
    product_type = _normalize_product_type(product_type)
    return PRODUCT_DESCRIPTIONS[product_type]


def template_names(product_type: str = DEFAULT_PRODUCT_TYPE) -> tuple[str, ...]:
    return tuple(_templates_for_product(product_type).keys())


def template_description(template: str, product_type: str = DEFAULT_PRODUCT_TYPE) -> str:
    return _template_spec(template, product_type).description


def _normalize_product_type(product_type: str | None) -> str:
    text = str(product_type or DEFAULT_PRODUCT_TYPE).strip().upper()
    if text not in PRODUCT_TYPES:
        raise ValueError(f"产品类型仅支持：{', '.join(PRODUCT_TYPES)}")
    return text


def _templates_for_product(product_type: str | None) -> dict[str, TemplateSpec]:
    return _PRODUCT_TEMPLATES[_normalize_product_type(product_type)]


def _template_spec(template: str, product_type: str | None = DEFAULT_PRODUCT_TYPE) -> TemplateSpec:
    templates = _templates_for_product(product_type)
    spec = templates.get(template)
    if spec is None:
        raise ValueError(f"标准模板仅支持：{', '.join(templates.keys())}")
    return spec


def _tol_wizard_rows(level: str, start_surface: int, end_surface: int) -> list[dict]:
    vals = _LEVEL_VALUES[level]
    rows = [
        ("半径", vals["半径"], "光圈"),
        ("厚度", vals["厚度"], "mm"),
        ("面倾斜X", vals["面倾斜"], "度"),
        ("面倾斜Y", vals["面倾斜"], "度"),
        ("元件偏心X", vals["元件偏心"], "mm"),
        ("元件偏心Y", vals["元件偏心"], "mm"),
        ("元件倾斜X", vals["元件倾斜"], "度"),
        ("元件倾斜Y", vals["元件倾斜"], "度"),
        ("面不规则", vals["面不规则"], "光圈"),
        ("折射率", vals["折射率"], "-"),
        ("阿贝%", vals["阿贝%"], "%"),
    ]
    return [
        {
            "启用": "Y",
            "公差类别": cat,
            "数值": value,
            "单位": unit,
            "起始面": start_surface,
            "结束面": end_surface,
            "跳过面": "",
        }
        for cat, value, unit in rows
    ]


def _resolve_params(params: dict, center_wave: int) -> dict:
    return {
        key: center_wave if value == "{center_wave}" else value
        for key, value in params.items()
    }


def _add_mfe(rows: list[dict], line_no: int, spec: OperandSpec,
             center_wave: int) -> None:
    params = _resolve_params(spec.params, center_wave)
    row = {
        "行号": line_no,
        "操作数": spec.op,
        "目标": spec.target,
        "权重": spec.weight,
        "注释": spec.label,
    }
    for i in range(1, 9):
        row[f"Param{i}"] = params.get(f"Param{i}", "")
    row["目标归一化视场"] = params.get("目标归一化视场", "")
    row["视场映射说明"] = "标准模板生成"
    row["归一化视场"] = row["目标归一化视场"]
    rows.append(row)


def _format_target_fields(fields: tuple[float, ...]) -> str:
    return ",".join(f"{float(field):g}" for field in fields)


def _mfe_and_report(template: str, center_wave: int,
                    product_type: str = DEFAULT_PRODUCT_TYPE) -> tuple[list[dict], list[dict]]:
    spec = _template_spec(template, product_type)

    mfe: list[dict] = []
    report: list[dict] = []
    line_no = 2
    for operand in spec.operands:
        _add_mfe(mfe, line_no, operand, center_wave)
        report.append({
            "启用": "Y",
            "标签": operand.label,
            "MF行号": line_no,
            "方向": operand.direction,
            "单位": operand.unit,
        })
        line_no += 1
    return mfe, report


def build_config(zmx_path: str, template: str = DEFAULT_TEMPLATE_NAME, level: str = "标准",
                 num_runs: int = 20, num_to_save: int = 0,
                 center_wave: int = 0, comp_mode: str = "无",
                 save_worst_best: bool = False,
                 product_type: str = DEFAULT_PRODUCT_TYPE) -> excel_io.Config:
    product_type = _normalize_product_type(product_type)
    template = template.strip() or DEFAULT_TEMPLATE_NAME
    level = level.strip() or "标准"
    spec = _template_spec(template, product_type)
    if level not in LEVEL_NAMES:
        raise ValueError(f"公差等级仅支持：{', '.join(LEVEL_NAMES)}")

    start_surface = 1
    end_surface = 0
    mfe, report = _mfe_and_report(template, center_wave, product_type)
    return excel_io.Config(
        tol_wizard=_tol_wizard_rows(level, start_surface, end_surface),
        tol_detail=[],
        mfe=mfe,
        report=report,
        run_params={
            "分析模式": "标准模板",
            "产品类型": product_type,
            "标准模板": template,
            "公差等级": level,
            "蒙特卡洛次数": int(num_runs),
            "保存数量": int(num_to_save),
            "统计分布": "正态",
            "补偿器模式": comp_mode,
            "TSC优化周期": 4,
            "中心波长号": int(center_wave),
            "后焦补偿面": "",
            "补偿Min": "",
            "补偿Max": "",
            "补偿线对": 17,
            "保存TSC": "Y",
            "保存WorstCase": "Y" if save_worst_best else "N",
            "保存BestCase": "Y" if save_worst_best else "N",
            "输出统计Excel": "Y",
            "输出直方图": "N",
            "启用中心指向偏移": "Y",
            "中心指向视场号": 1,
            "启用焦距偏移百分比": "Y",
            "启用视场映射": "Y",
            "视场插入策略": "自动插入",
            "视场匹配阈值": 0.001,
            "目标归一化视场": _format_target_fields(spec.target_fields),
            "目标视场来源策略": "自动推断",
        },
    )


def _rewrite_sheet(ws, header: list[str], rows: list[dict]) -> None:
    ws.delete_rows(2, ws.max_row)
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(header, start=1):
            ws.cell(row=r, column=c, value=row.get(key))


def write_config_excel(path: str, cfg: excel_io.Config, overwrite: bool = True) -> str:
    if os.path.exists(path) and not overwrite:
        raise FileExistsError(path)
    excel_io.generate_template(path, overwrite=True)
    wb = load_workbook(path)
    _rewrite_sheet(wb["输入_公差向导"], excel_io._TOL_WIZARD_HDR, cfg.tol_wizard)
    _rewrite_sheet(wb["输入_公差明细"], excel_io._TOL_DETAIL_HDR, cfg.tol_detail)
    _rewrite_sheet(wb["输入_评价函数"], excel_io._MFE_HDR, cfg.mfe)
    _rewrite_sheet(wb["输入_REPORT"], excel_io._REPORT_HDR, cfg.report)
    run_rows = [{"参数键": k, "值": v, "备注": "标准模板生成"} for k, v in cfg.run_params.items()]
    _rewrite_sheet(wb["输入_运行参数"], excel_io._RUN_HDR, run_rows)
    wb.save(path)
    return path


def default_config_path(zmx_path: str, outdir: str | None = None) -> str:
    parent = os.path.abspath(outdir) if outdir else os.path.dirname(os.path.abspath(zmx_path))
    base = os.path.splitext(os.path.basename(zmx_path))[0]
    safe = re.sub(r'[^0-9A-Za-z_\-\u4e00-\u9fff]+', "_", base).strip("._") or "lens"
    return os.path.join(parent, f"{safe}_标准模板配置.xlsx")


def make_temp_config(zmx_path: str, outdir: str | None, template: str, level: str,
                     num_runs: int, num_to_save: int, center_wave: int,
                     comp_mode: str, save_worst_best: bool = False,
                     product_type: str = DEFAULT_PRODUCT_TYPE) -> str:
    parent = os.path.abspath(outdir) if outdir else os.path.dirname(os.path.abspath(zmx_path))
    os.makedirs(parent, exist_ok=True)
    base = os.path.splitext(os.path.basename(zmx_path))[0]
    safe = re.sub(r'[^0-9A-Za-z_\-\u4e00-\u9fff]+', "_", base).strip("._") or "lens"
    path = os.path.join(parent, f"{safe}_标准模板配置_{uuid.uuid4().hex[:8]}.xlsx")
    cfg = build_config(zmx_path, template=template, level=level,
                       num_runs=num_runs, num_to_save=num_to_save,
                       center_wave=center_wave, comp_mode=comp_mode,
                       save_worst_best=save_worst_best,
                       product_type=product_type)
    return write_config_excel(path, cfg, overwrite=False)
