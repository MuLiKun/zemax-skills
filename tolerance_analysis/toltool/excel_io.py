"""excel_io.py —— 配置 Excel 的读写与模板生成。

职责（需求文档 §4、§7）：
- generate_template(): 生成带「示例_*」(只读参考) 与「输入_*」(用户填写) 双 sheet 的配置模板。
- read_config(): 读取「输入_*」各 sheet，返回结构化配置。
- 约定：sheet 名前缀区分示例与输入；表头第 1 行；数据第 2 行起；空行跳过；# 开头注释。
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


_HDR_FILL = PatternFill("solid", fgColor="DDEBF7")
_EX_FILL = PatternFill("solid", fgColor="FFF2CC")
_HDR_FONT = Font(bold=True)
_VALIDATION_MIN_ROWS = 300


# ---------------------------------------------------------------------------
# 各 sheet 的表头与示例数据定义（需求文档 §4.3 ~ §4.7）
# ---------------------------------------------------------------------------

_TOL_WIZARD_HDR = ["启用", "公差类别", "数值", "单位", "起始面", "结束面", "跳过面"]
_TOL_WIZARD_EX = [
    ["Y", "半径", 3, "光圈", 3, 13, 9],
    ["Y", "厚度", 0.03, "mm", 3, 13, 9],
    ["Y", "面偏心X", 0.02, "mm", 3, 13, 9],
    ["Y", "面偏心Y", 0.02, "mm", 3, 13, 9],
    ["Y", "面倾斜X", 0.2, "度", 3, 13, 9],
    ["Y", "面倾斜Y", 0.2, "度", 3, 13, 9],
    ["Y", "元件偏心X", 0.02, "mm", 3, 13, 9],
    ["Y", "元件偏心Y", 0.02, "mm", 3, 13, 9],
    ["Y", "元件倾斜X", 0.2, "度", 3, 13, 9],
    ["Y", "元件倾斜Y", 0.2, "度", 3, 13, 9],
    ["Y", "面不规则", 1, "光圈", 3, 13, 9],
    ["N", "Zernike不规则度", 1, "光圈", 3, 13, 9],
    ["Y", "折射率", 0.0005, "-", 3, 13, 9],
    ["Y", "阿贝%", 1, "%", 3, 13, 9],
]

_TOL_DETAIL_HDR = ["操作数", "面1", "面2", "Min", "Max", "注释"]
_TOL_DETAIL_LEGACY_HDR = ["操作", "操作数", "面1", "面2", "Min", "Max", "注释"]
_TOL_DETAIL_EX = [
    ["TFRN", 5, 5, -3, 3, "第5面半径3环"],
    ["TTHI", 6, 6, -0.05, 0.05, "第6面厚度公差"],
    ["TIRR", 7, 7, -1, 1, "第7面不规则"],
]

_MFE_HDR = ["行号", "操作数",
            "Param1", "Param2", "Param3", "Param4",
            "Param5", "Param6", "Param7", "Param8",
            "目标", "权重", "注释", "目标归一化视场", "视场映射说明"]
_MFE_EX = [
    [2, "RSCE", 3, 2, 0, 0.0, "", "", "", "", 0, 1, "点列 F0；RSCE 的 Param4 即归一化视场", "", "空则从 Param4 推断"],
    [3, "RSCE", 3, 2, 0, 0.5, "", "", "", "", 0, 1, "点列 F0.5", "", "空则从 Param4 推断"],
    [4, "RSCE", 3, 2, 0, 0.9, "", "", "", "", 0, 1, "点列 F0.9", "", "空则从 Param4 推断"],
    [5, "GENC", 3, 2, 1, 0.95, 1, 0, 0, "", 0, 1, "GENC95；Param3 可直接从 Zemax MF 复制", "", "空则从 Param3 原视场号推断"],
    [6, "GENC", 3, 2, 5, 0.95, 1, 0, 0, "", 0, 1, "GENC95；示例从原视场号推断", "", "专家可填 0.7/0.9 等覆盖"],
    [7, "GENC", 3, 2, 7, 0.95, 1, 0, 0, "", 0, 1, "GENC95；示例从原视场号推断", "", "映射结果写入 field_mapping.txt"],
    [8, "GMTT", 3, 2, 1, 34, 0, 0, "", "", 0, 1, "几何MTF子午；Param3 可从 Zemax MF 复制", "", "空则从 Param3 原视场号推断"],
    [9, "GMTS", 3, 2, 1, 34, 0, 0, "", "", 0, 1, "几何MTF弧矢；Param3 可从 Zemax MF 复制", "", "空则从 Param3 原视场号推断"],
    [10, "GMTA", 3, 2, 1, 34, 0, 0, "", "", 0, 1, "几何MTF平均；Param3 可从 Zemax MF 复制", "", "空则从 Param3 原视场号推断"],
]

_REPORT_HDR = ["启用", "标签", "MF行号", "方向", "单位"]
_REPORT_EX = [
    ["Y", "SPOT_Hy0.0", 2, "小", "mm"],
    ["Y", "SPOT_Hy0.5", 3, "小", "mm"],
    ["Y", "SPOT_Hy0.9", 4, "小", "mm"],
    ["Y", "GENC95_F1", 5, "小", "um"],
    ["Y", "GMTFT_F1", 8, "大", "-"],
    ["Y", "GMTFS_F1", 9, "大", "-"],
]

_RUN_HDR = ["参数键", "值", "备注"]
_RUN_EX = [
    ["蒙特卡洛次数", 200, "NumberOfRuns"],
    ["保存数量", 10, "NumberToSave 保存最差前N个"],
    ["统计分布", "正态", "正态/均匀/抛物线"],
    ["补偿器模式", "无", "无/全部优化DLS/全部优化OD"],
    ["TSC优化周期", 4, "TSC 内 OPTIMIZE n"],
    ["中心波长号", 2, "评估用 Wave 编号"],
    ["后焦补偿面", "", "COMP 补偿面号；留空=不加补偿器"],
    ["补偿Min", "", "补偿下限；留空=Zemax 自由调整"],
    ["补偿Max", "", "补偿上限；留空=Zemax 自由调整"],
    ["补偿线对", 34, "补偿专用 MF 的 GMTA 频率 lp/mm"],
    ["保存TSC", "Y", "始终保存"],
    ["保存WorstCase", "Y", "保存最差案例文件"],
    ["保存BestCase", "Y", "保存最佳案例文件"],
    ["输出统计Excel", "Y", "Y=导出ZTD统计Excel，含百分位与Cpk1.33规格限"],
    ["输出直方图", "N", "本期预留，默认关"],
    ["启用视场映射", "N", "Y=启用后台视场映射；默认关"],
    ["视场插入策略", "禁用", "禁用/自动插入；自动插入只修改 tol 工作副本"],
    ["视场匹配阈值", 0.05, "目标归一化视场与最近已有视场差值大于该值时视为缺失"],
    ["目标归一化视场", "0,-0.25,0.25,-0.5,0.5,-0.7,0.7,-0.9,0.9,-1,1", "标准目标视场序列；后台推断时吸附到最近目标"],
    ["目标视场来源策略", "自动推断", "自动推断=优先目标列，空则从RSCE Param4/其他操作数Param3推断；仅显式=只用目标列，不自动反推"],
]

_INTRO_LINES = [
    "Zemax 公差分析自动化程序 —— 配置说明",
    "",
    "1. 本工作簿用 sheet 名前缀区分：",
    "   示例_*  ：只读参考模板，请勿删除，程序忽略其内容。",
    "   输入_*  ：用户实际填写，程序只读取这些 sheet。",
    "2. 每个 sheet 第 1 行是表头；数据从第 2 行开始；空行跳过；# 开头的行视为注释。",
    "3. 公差表两种方式可叠加：先用 输入_公差向导 批量生成默认，",
    "   再用 输入_公差明细 逐行追加公差；该表可直接粘贴 Zemax GUI 中间数据行。",
    "   旧模板中若仍保留“操作”列，程序仍兼容追加/覆盖/删除。",
    "4. 评价函数 sheet 的 行号 与 REPORT sheet 的 MF行号 一一对应。",
    "5. 每个启用的 REPORT 行将在蒙特卡洛结果中成为一个独立分项列。",
    "6. 运行参数中 保存TSC 始终为 Y；保存WorstCase/BestCase 由用户选择 Y/N。",
    "7. 后焦补偿面留空=不加 TDE 补偿器；填面号则在该面厚度加 COMP，",
    "   补偿Min/Max 留空时由 Zemax 自由调整，补偿线对为补偿专用 MF 的 GMTA 频率。",
]


_SHEET_DEFS = [
    ("公差向导", _TOL_WIZARD_HDR, _TOL_WIZARD_EX),
    ("公差明细", _TOL_DETAIL_HDR, _TOL_DETAIL_EX),
    ("评价函数", _MFE_HDR, _MFE_EX),
    ("REPORT", _REPORT_HDR, _REPORT_EX),
    ("运行参数", _RUN_HDR, _RUN_EX),
]


def _write_table(ws: Worksheet, header: list[str], rows: list[list[Any]],
                 highlight: bool) -> None:
    for c, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if highlight:
                cell.fill = _EX_FILL
    for c, name in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = \
            max(10, len(str(name)) + 4)


def _add_list_validation(ws: Worksheet, target: str, values: list[str], rows: int | None = None) -> None:
    formula = '"' + ','.join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    if ":" in target:
        cell_range = target
    else:
        min_rows = _VALIDATION_MIN_ROWS if rows is None else int(rows)
        end_row = max(ws.max_row, min_rows, 2)
        cell_range = f"{target}2:{target}{end_row}"
    dv.add(cell_range)


def _apply_template_validations(ws: Worksheet, name: str) -> None:
    if name == "公差向导":
        _add_list_validation(ws, "A", ["Y", "N"])
        _add_list_validation(ws, "B", [
            "半径", "曲率半径", "厚度", "面偏心X", "面偏心Y",
            "面倾斜X", "面倾斜Y", "元件偏心X", "元件偏心Y",
            "元件倾斜X", "元件倾斜Y", "面不规则",
            "Zernike不规则度", "折射率", "阿贝%",
        ])
        for row in range(2, ws.max_row + 1):
            cat = str(ws.cell(row=row, column=2).value or "").strip()
            if cat in ("半径", "曲率半径"):
                _add_list_validation(ws, f"D{row}:D{row}", ["光圈", "毫米", "百分比"])
            elif cat in ("面倾斜X", "面倾斜Y"):
                _add_list_validation(ws, f"D{row}:D{row}", ["度", "毫米"])


def generate_template(path: str, overwrite: bool = False) -> str:
    """生成带示例与输入双 sheet 的配置模板。返回写入路径。"""
    if os.path.exists(path) and not overwrite:
        raise FileExistsError(f"{path} 已存在，使用 overwrite=True 覆盖。")

    wb = Workbook()
    intro = wb.active
    intro.title = "说明"
    for r, line in enumerate(_INTRO_LINES, start=1):
        c = intro.cell(row=r, column=1, value=line)
        if r == 1:
            c.font = Font(bold=True, size=13)
    intro.column_dimensions["A"].width = 70

    for name, header, ex in _SHEET_DEFS:
        ws_ex = wb.create_sheet(f"示例_{name}")
        _write_table(ws_ex, header, ex, highlight=True)
        _apply_template_validations(ws_ex, name)
        ws_in = wb.create_sheet(f"输入_{name}")
        prefill = ex if name in ("公差向导", "评价函数", "REPORT", "运行参数") else []
        _write_table(ws_in, header, prefill, highlight=False)
        _apply_template_validations(ws_in, name)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 读取配置
# ---------------------------------------------------------------------------

@dataclass
class Config:
    tol_wizard: list[dict] = field(default_factory=list)
    tol_detail: list[dict] = field(default_factory=list)
    mfe: list[dict] = field(default_factory=list)
    report: list[dict] = field(default_factory=list)
    run_params: dict = field(default_factory=dict)


def _read_sheet_rows(ws: Worksheet, header: list[str]) -> list[dict]:
    actual_cols = sum(
        1 for i in range(1, len(header) + 1)
        if str(ws.cell(row=1, column=i).value or "").strip())
    if actual_cols < len(header):
        raise ValueError(
            f"工作表 {ws.title!r} 的表头列数不足：期望 {len(header)} 列"
            f"（{', '.join(header)}），实际仅 {actual_cols} 列。"
            f"请使用最新模板或补全表头。")
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or v == "" for v in row):
            continue
        first = row[0]
        if isinstance(first, str) and first.strip().startswith("#"):
            continue
        rows.append({header[i]: row[i] if i < len(row) else None
                     for i in range(len(header))})
    return rows


def _read_tol_detail_rows(ws: Worksheet) -> list[dict]:
    first_header = str(ws.cell(row=1, column=1).value or "").strip()
    if first_header == "操作":
        return _read_sheet_rows(ws, _TOL_DETAIL_LEGACY_HDR)
    rows = _read_sheet_rows(ws, _TOL_DETAIL_HDR)
    for row in rows:
        row["操作"] = "追加"
    return rows


def _read_mfe_rows(ws: Worksheet) -> list[dict]:
    actual = [str(ws.cell(row=1, column=i).value or "").strip()
              for i in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if values is None or all(v is None or v == "" for v in values):
            continue
        first = values[0]
        if isinstance(first, str) and first.strip().startswith("#"):
            continue
        row = {actual[i]: values[i] if i < len(values) else None
               for i in range(len(actual)) if actual[i]}
        row.setdefault("目标归一化视场", row.get("归一化视场"))
        row.setdefault("归一化视场", row.get("目标归一化视场"))
        row.setdefault("视场映射说明", None)
        # 持久列以 _MFE_HDR 为准（含“目标归一化视场”）；
        # “归一化视场”仅作运行期内存兼容键，不写回 Excel（见 write_mapped_config）。
        rows.append({h: row.get(h) for h in _MFE_HDR} | {
            "归一化视场": row.get("归一化视场"),
        })
    return rows


def read_config(path: str) -> Config:
    wb = load_workbook(path, data_only=True)
    cfg = Config()

    def sheet(name: str):
        sn = f"输入_{name}"
        if sn not in wb.sheetnames:
            raise KeyError(f"缺少 sheet: {sn}")
        return wb[sn]

    cfg.tol_wizard = _read_sheet_rows(sheet("公差向导"), _TOL_WIZARD_HDR)
    cfg.tol_detail = _read_tol_detail_rows(sheet("公差明细"))
    cfg.mfe = _read_mfe_rows(sheet("评价函数"))
    cfg.report = _read_sheet_rows(sheet("REPORT"), _REPORT_HDR)
    for row in _read_sheet_rows(sheet("运行参数"), _RUN_HDR):
        key = row.get("参数键")
        if key:
            cfg.run_params[str(key).strip()] = row.get("值")
    return cfg


def _rewrite_sheet(ws: Worksheet, header: list[str], rows: list[dict]) -> None:
    ws.delete_rows(1, ws.max_row)
    _write_table(ws, header, [[row.get(h) for h in header] for row in rows], highlight=False)


def write_config_snapshot(source_path: str, target_path: str, cfg: Config) -> str:
    # 注：此处 load_workbook 不带 data_only=True，保留源文件公式以维持快照可读性；
    # read_config 用 data_only=True 读取计算值。若源含公式，二者展示值可能不同，
    # 但本快照仅供人工核对最终运行配置，不参与计算，差异可接受。
    wb = load_workbook(source_path)
    if "输入_公差向导" in wb.sheetnames:
        _rewrite_sheet(wb["输入_公差向导"], _TOL_WIZARD_HDR, cfg.tol_wizard)
    if "输入_公差明细" in wb.sheetnames:
        _rewrite_sheet(wb["输入_公差明细"], _TOL_DETAIL_HDR, cfg.tol_detail)
    if "输入_评价函数" in wb.sheetnames:
        _rewrite_sheet(wb["输入_评价函数"], _MFE_HDR, cfg.mfe)
    if "输入_REPORT" in wb.sheetnames:
        _rewrite_sheet(wb["输入_REPORT"], _REPORT_HDR, cfg.report)
    if "输入_运行参数" in wb.sheetnames:
        run_rows = [{"参数键": k, "值": v, "备注": "最终运行配置"} for k, v in cfg.run_params.items()]
        _rewrite_sheet(wb["输入_运行参数"], _RUN_HDR, run_rows)
    wb.save(target_path)
    return target_path


def write_mapped_config(source_path: str, target_path: str, cfg: Config) -> str:
    return write_config_snapshot(source_path, target_path, cfg)
